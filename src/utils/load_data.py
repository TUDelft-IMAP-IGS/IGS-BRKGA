import pandas as pd

import sys
import json

from itertools import chain
from openpyxl import load_workbook, Workbook
from datetime import date, datetime
from loguru import logger
from werkzeug.datastructures import FileStorage
from io import BytesIO

from components.vessel import Vessel
from components.activity import Activity


def get_data_sheet(wb: Workbook, sheet_name: str, max_col: int) -> list[dict]:
    """Get relevant data from specific sheet in the excel workbook."""
    try:
        # open sheet
        ws = wb[sheet_name]
    except KeyError:
        logger.exception(f"Sheet {sheet_name} is not found in excel.")
        sys.exit(1)
    except Exception as e:
        logger.exception(f"Unable to open sheet: {sheet_name}")
        sys.exit(1)

    # get data from core portion of the spreadsheet (columns: 2 - max_col, rows: 6 - ...)
    # ignoring null rows
    data = [row for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=2, max_col=max_col, values_only=True) if any(cell is not None for cell in row) and (row[0] is not None)]

    # convert data into a dictionary, with the first row corresponding to the keys
    data_dict = [dict(zip(data[0], row)) for row in data[1:]]

    return data_dict


def get_coordinates(data_locations: list[dict], location_name: str) -> tuple | None:
    """Get coordinates from corresponding location list."""

    for entry in data_locations:
        # find corresponding entry
        if entry.get("location_name") == location_name:
            # return coordinates if found
            return (entry.get("longitude"), entry.get("latitude"))

    return None


def get_coordinates_list(data_locations: list[dict], location_names: str) -> tuple[list, list]:
    """Get list of coordinates and names for a list of multiple possible locations."""
    locations_coords = []
    locations_names = []

    if location_names.lower() == "all":
        # all possible locations allowed
        locations_coords = [(row["longitude"], row["latitude"]) for row in data_locations]
        locations_names = [row["location_name"] for row in data_locations]
    else:
        # convert input string to list
        locations_names_list = [x.strip() for x in location_names.split(",")]

        for name in locations_names_list:
            # get corresponding coordinates
            coords = get_coordinates(data_locations, name)

            if coords is not None:
                locations_coords.append(coords)
                locations_names.append(name)

    return locations_coords, locations_names


def get_possible_vessel_configs(vessels: list, vessel_groups: list, configs_text: str, counts_text: str) -> list | None:
    """Setup possible vessel configurations based on input requirements."""
    # initialise possible configs
    possible_configs = [[]]

    # lists of vessels and groups
    vessels_list = [v["vessel_name"] for v in vessels]
    groups_dict = {g.get("group_name", ""): [x.strip() for x in g.get("list_vessels", "").split(",")] for g in vessel_groups}

    # convert input strings to list
    configs_list = [x.strip() for x in configs_text.split(",")]
    counts_list = [int(x.strip()) for x in str(counts_text).split(",")]

    if len(configs_list) != len(counts_list):
        logger.warning("Configuration list and list of counts have different lengths.")
        return None

    for i in range(len(configs_list)):
        if configs_list[i] in vessels_list:
            # single vessel
            possible_configs[0].append({"vessels": [configs_list[i]], "count": 1})
        elif configs_list[i] in groups_dict.keys():
            # group of vessels
            corresp_vessels = groups_dict[configs_list[i]]
            possible_configs[0].append({"vessels": corresp_vessels, "count": counts_list[i]})
        else:
            logger.warning("Invalid entry in configuration list.")
            return None

    return possible_configs


def date_to_str(val: datetime | str | None) -> str | None:
    """Convert a datetime to dd/mm/YYYY string (leaving strings or None unchanged.)"""

    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    else:
        return val


def load_input_json(input_data: str | dict, date_origin: date):
    """Read string input and generate vessels and activities."""
    if isinstance(input_data, str):
        # convert string to dictionary
        logger.info("Convert input to dict.")
        input_data = json.loads(input_data)

    if isinstance(input_data, dict):
        # get vessels and activities separately
        data_vessels = input_data.get("vessels", [])
        data_activites = input_data.get("activities", [])
        data_current_locations = input_data.get("current_locations", [])

        # generate vessel and activity objects
        # logger.debug("Generating vessels.")
        vessels = [Vessel(row) for row in data_vessels]
        # logger.debug("Generating activities.")
        activities = [Activity(row, date_origin) for row in data_activites if row.get("is_valid")]

        # get list of vessels which are already assigned an activity for today (ignore current location)
        allowed_lists = [a.allowed_vessels for a in activities if a.is_current_activity]
        all_vessels = list(chain.from_iterable(allowed_lists))
        unique_vessels = list(dict.fromkeys(all_vessels))

        # filter for current location, solver will handle this logic
        current_locations = [Activity(row, date_origin) for row in data_current_locations if row["target_vessel"] not in unique_vessels]

        logger.info(f"Created {len(vessels)} vessels, {len(activities)} projects and maintenance activities and {len(current_locations)} current locations.")

        activities += current_locations

        return vessels, activities
    else:
        logger.error(f"Input json does not have dict formatting: {type(input_data)}")
        raise Exception(f"Input json does not have dict formatting: {type(input_data)}")


def load_excel_from_filestorage(file: FileStorage):
    """Read excel file uploaded from Flask directly from memory."""

    # go to beginning of stream
    file.stream.seek(0)

    # load workbook using openpyxl
    workbook = load_workbook(filename=BytesIO(file.read()), data_only=True, read_only=True)

    return workbook


def load_input_excel(file: FileStorage | str, date_origin: date) -> tuple[list, list]:
    """Read excel from input and generate vessels, projects and maintenance objects."""

    file_name = file

    try:
        # load entire excel file
        if isinstance(file, FileStorage):
            file_name = file.filename
            # logger.info(f"Opening excel spreadsheet: {file_name}")
            wb = load_excel_from_filestorage(file)
        else:
            # logger.info(f"Opening excel spreadsheet: {file_name}")
            wb = load_workbook(file, data_only=True, read_only=True)

    except FileNotFoundError:
        logger.exception(f"File not found: {file_name}")
        sys.exit(1)
    except PermissionError:
        logger.exception(f"permission denied: {file_name}")
        sys.exit(1)
    except Exception as e:
        logger.exception(f"Failed to open file: {file_name}\n{e}")
        sys.exit(1)

    # extract each table from each sheet
    data_vessels = get_data_sheet(wb, "vessels", 12)
    data_vessel_groups = get_data_sheet(wb, "vessel_groups", 5)
    data_projects = get_data_sheet(wb, "projects", 17)

    data_project_locations = get_data_sheet(wb, "project_locations", 5)
    data_maintenance = get_data_sheet(wb, "maintenance", 9)
    data_maintenance_locations = get_data_sheet(wb, "maintenance_locations", 5)

    # update project inputs
    for i in range(len(data_projects)):
        # start and end locations to coordinates
        data_projects[i]["start_location"] = get_coordinates(data_project_locations, data_projects[i].get("start_location_name", ""))
        data_projects[i]["end_location"] = get_coordinates(data_project_locations, data_projects[i].get("end_location_name", ""))

        # set start and end date to be string
        data_projects[i]["actual_start"] = date_to_str(data_projects[i]["actual_start"])
        data_projects[i]["begin_start_window"] = date_to_str(data_projects[i]["begin_start_window"])
        data_projects[i]["end_start_window"] = date_to_str(data_projects[i]["end_start_window"])

        if data_projects[i]["start_location"] is None:
            logger.warning(f"Invalid location provided for {data_projects[i]['activity_name']}: {data_projects[i]['start_location']}. Project skipped.")
            continue

        if data_projects[i]["end_location"] is None:
            logger.warning(f"Invalid location provided for {data_projects[i]['activity_name']}: {data_projects[i]['end_location']}. Project skipped.")
            continue

        # vessel configuration
        data_projects[i]["required_vessel_config"] = get_possible_vessel_configs(
            data_vessels,
            data_vessel_groups,
            data_projects[i].get("vessel_configs", ""),
            data_projects[i].get("vessel_counts", ""),
        )

        if data_projects[i]["required_vessel_config"] is None:
            logger.warning(f"Failed to generate required vessel config for {data_projects[i]['activity_name']}. Project skipped.")
            continue

        # correct activity type
        data_projects[i]["activity_type"] = "tow"

        # flag entry as valid
        data_projects[i]["is_valid"] = True

    # update maintenance inputs
    for i in range(len(data_maintenance)):
        # possible locations
        (
            data_maintenance[i]["allowed_locations"],
            data_maintenance[i]["allowed_locations_names"],
        ) = get_coordinates_list(
            data_maintenance_locations,
            data_maintenance[i].get("possible_locations", ""),
        )

        # set start and end date to be string
        data_maintenance[i]["actual_start"] = date_to_str(data_maintenance[i]["actual_start"])
        data_maintenance[i]["begin_start_window"] = date_to_str(data_maintenance[i]["begin_start_window"])
        data_maintenance[i]["end_start_window"] = date_to_str(data_maintenance[i]["end_start_window"])

        # correct activity type
        data_maintenance[i]["activity_type"] = "maintenance"

        # flag entry as valid
        data_maintenance[i]["is_valid"] = True

    vessels = data_vessels
    projects = [row for row in data_projects if row.get("is_valid")]
    maintenance = [row for row in data_maintenance if row.get("is_valid")]
    activities = projects + maintenance

    # logger.info(f"Created {len(vessels)} vessels, {len(projects)} projects and {len(maintenance)} maintenance activities.")

    return vessels, activities


def load_input_excel_to_json(file: FileStorage | str, date_origin: date) -> tuple[dict, int, int]:
    """Convert input excel to jsons for vessels and activities."""

    vessels, activities = load_input_excel(file, date_origin)

    input_dict = {"vessels": vessels, "activities": activities}

    return input_dict, len(vessels), len(activities)

import pandas as pd
import numpy as np
import random
import json
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# METADATA TEMPLATES (Rows 1-6 for every sheet)
# =============================================================================
METADATA = {
    "vessels": [
        [],
        ["Table Name", "Vessels"],
        ["Description", "Vessel Name", "Vessel IMO", "Sailing Speed (ECO)\n[kn]", "Sailing Speed (CRUISE)\n[kn]", "Sailing Speed (MAX)\n[kn]", "Fuel Consumed (ECO)\n[MT/day]", "Fuel Consume (CRUISE)\n[MT/day]", "Fuel Consumed (MAX)\n[MT/day]", "Mobilization Day Rate\nNorm Tariff - 2025\n[USD]"],
        ["Formatting", "text", "number", "number", "number", "number", "number", "number", "number", "number"],
        ["", "", "", "", "", "", "", "", "", "", "INPUT VALIDATION"],
        ["row_id", "vessel_name", "vessel_IMO", "sailing_speed_eco", "sailing_speed_cruise", "sailing_speed_max", "fuel_consumed_eco", "fuel_consumed_cruise", "fuel_consumed_max", "day_rate_mob", "is_entry", "is_valid_speeds", "is_valid_fuel", "is_valid_rate", "is_valid_name", "is_valid_IMO"]
    ],
    "vessel_groups": [
        [],
        ["Table Name", "Vessel Groups"],
        ["Description", "Group Name", "Group Description", "Included Vessels"],
        ["Formatting", "text", "text", "vessel 1, vessel 2, …"],
        ["", "", "", "", "INPUT VALIDATION"],
        ["row_id", "group_name", "group_description", "list_vessels", "is_entry", "is_valid_name", "is_valid_vessels"]
    ],
    "projects": [
        [],
        ["Table Name", "Projects"],
        ["Description", "BIMS Project Code", "Project Name", "Actual Start Date", "Begin Start Window", "End Start Window\n*only include this date if you would like to have variable start days", "Project Duration\nFIXED\n[days]", "Project Duration\nOPTIONS\n[days]", "Project Duration\nOperational Expectation\n[days]", "Start Location", "End Location", "Route Description", "Vessel Configuration", "Corresponding Number of Vessels"],
        ["Formatting", "text", "text", "dd/mm/yyyy", "dd/mm/yyyy", "dd/mm/yyyy", "integer", "integer", "integer", "string", "string", "text", "list (with commas)", "list (with commas)"],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "INPUT VALIDATION"],
        ["row_id", "project_code", "activity_name", "actual_start", "begin_start_window", "end_start_window", "duration_fixed", "duration_options", "duration_opex", "start_location_name", "end_location_name", "route_description", "vessel_configs", "vessel_counts", "is_entry", "is_valid_name", "is_valid_dates", "is_valid_locations", "is_valid_durations", "is_valid_configs", "is_valid_counts"]
    ],
    "project_locations": [
        [],
        ["Table Name", "Project Locations"],
        ["Description", "Location/Port Name", "Longitude", "Latitude"],
        ["Formatting", "text", "number", "number"],
        ["", "", "", "", "INPUT VALIDATION"],
        ["row_id", "location_name", "longitude", "latitude", "is_entry", "is_valid_name", "is_valid_longitude", "is_valid_latitude"]
    ],
    "maintenance": [
        [],
        ["Table Name", "Maintenance"],
        ["Description", "Maintenance Name", "Actual Start Date", "Begin Start Window", "End Start Window", "Maintenance Duration\n[days]", "Target Vessel", "Possible Locations"],
        ["Formatting", "text", "dd/mm/yyyy", "dd/mm/yyyy", "dd/mm/yyyy", "integer", "text", "text"],
        ["", "", "", "", "", "", "", "", "", "", "INPUT VALIDATION"],
        ["row_id", "activity_name", "actual_start", "begin_start_window", "end_start_window", "duration", "target_vessel", "possible_locations", "is_entry", "is_valid_name", "is_valid_dates", "is_valid_durations", "is_valid_vessel", "is_valid_locations"]
    ],
    "maintenance_locations": [
        [],
        ["Table Name", "Maintenance Locations"],
        ["Description", "Location/Port Name", "Longitude", "Latitude"],
        ["Formatting", "text", "number", "number"],
        ["", "", "", "", "INPUT VALIDATION"],
        ["row_id", "location_name", "longitude", "latitude", "is_entry", "is_valid_name", "is_valid_longitude", "is_valid_latitude"]
    ]
}

# =============================================================================
# GEOGRAPHIC CLUSTERS
# =============================================================================
GEO_CLUSTERS = {
    "north_sea": {
        "lon": (2.0, 8.0),
        "lat": (54.0, 62.0),
        "port_names": [
            "Rotterdam", "Amsterdam", "Esbjerg", "Aberdeen", "Stavanger",
            "Bergen", "Kristiansand", "Groningen", "Immingham", "Dundee",
        ],
    },
    "gulf_of_mexico": {
        "lon": (-97.0, -84.0),
        "lat": (18.0, 30.0),
        "port_names": [
            "Houston", "Veracruz", "Tampa", "New Orleans", "Corpus Christi",
            "Tampico", "Mobile", "Brownsville", "Galveston", "Pensacola",
        ],
    },
    "southeast_asia": {
        "lon": (100.0, 115.0),
        "lat": (-5.0, 10.0),
        "port_names": [
            "Singapore", "Kuala Lumpur", "Jakarta", "Batam", "Johor Bahru",
            "Kota Kinabalu", "Pontianak", "Palembang", "Belawan", "Dumai",
        ],
    },
}

DEFAULT_CLUSTER = "north_sea"

# =============================================================================
# PARAMETER GUARD-RAILS
# =============================================================================
WINDOW_LOOSE_MIN_DAYS = 14
WINDOW_LOOSE_MAX_DAYS = 90
WINDOW_TIGHT_MIN_DAYS = 14
WINDOW_TIGHT_MAX_DAYS = 30

DURATION_MIN_DAYS = 5
DURATION_MAX_DAYS = 80

MAINT_DURATION_MIN = 5
MAINT_DURATION_MAX = 14

HORIZON_DAYS = 270

MAINT_VESSEL_FRACTION = 0.40


# =============================================================================
# VESSEL NAME HELPER
# =============================================================================

def _vessel_name(i: int) -> str:
    """
    Single source of truth for vessel naming.
    Both generate_vessels() and generate_current_locations() must use this
    function so that target_vessel values in the JSON always match the
    vessel_name values in the Excel file exactly.
    """
    return f"Vessel {i + 1}"


# =============================================================================
# DATA GENERATORS
# =============================================================================

def _cluster_coords(cluster_name: str) -> tuple[tuple, tuple, list]:
    c = GEO_CLUSTERS.get(cluster_name, GEO_CLUSTERS[DEFAULT_CLUSTER])
    return c["lon"], c["lat"], c["port_names"]


def generate_locations(num_locations: int, cluster: str = DEFAULT_CLUSTER) -> pd.DataFrame:
    lon_range, lat_range, port_names = _cluster_coords(cluster)
    locations = []

    for i in range(num_locations):
        loc_name = port_names[i % len(port_names)]
        if i >= len(port_names):
            loc_name = f"{loc_name}_{i}"

        locations.append({
            "row_id": i + 1,
            "location_name": loc_name,
            "longitude": round(random.uniform(*lon_range), 3),
            "latitude": round(random.uniform(*lat_range), 3),
            "is_entry": True,
            "is_valid_name": True,
            "is_valid_longitude": True,
            "is_valid_latitude": True,
        })
    return pd.DataFrame(locations)


def generate_vessels(num_vessels: int) -> pd.DataFrame:
    vessels = []
    for i in range(num_vessels):
        vessels.append({
            "row_id": i + 1,
            # FIX: use shared helper so name is always consistent with JSON output
            "vessel_name": _vessel_name(i),
            "vessel_IMO": random.randint(9000000, 9999999),
            "sailing_speed_eco": random.randint(70, 100) * 0.1,
            "sailing_speed_cruise": random.randint(110, 150) * 0.1,
            "sailing_speed_max": random.randrange(160, 200) * 0.1,
            "fuel_consumed_eco": random.randint(140, 160) * 0.1,
            "fuel_consumed_cruise": random.randint(250, 400) * 0.1,
            "fuel_consumed_max": random.randint(650, 800) * 0.1,
            "day_rate_mob": random.choice([22500, 31650, 45000]),
            "is_entry": True,
            "is_valid_speeds": True,
            "is_valid_fuel": True,
            "is_valid_rate": True,
            "is_valid_name": True,
            "is_valid_IMO": True,
        })
    return pd.DataFrame(vessels)


def generate_vessel_groups(vessels_df: pd.DataFrame, num_vessels: int) -> pd.DataFrame:
    MIN_CLASS_SIZE = 2

    groups = []
    vessel_names = vessels_df["vessel_name"].tolist()

    groups.append({
        "row_id": 1,
        "group_name": "ALL",
        "group_description": "All Vessels",
        "list_vessels": "*",
        "is_entry": True,
        "is_valid_name": True,
        "is_valid_vessels": True,
    })

    num_classes = max(1, num_vessels // MIN_CLASS_SIZE)
    num_classes = min(num_classes, 3)
    chunk_size = num_vessels // num_classes

    for i in range(num_classes):
        start_idx = i * chunk_size
        end_idx = start_idx + chunk_size if i < num_classes - 1 else num_vessels
        class_vessels = vessel_names[start_idx:end_idx]

        groups.append({
            "row_id": i + 2,
            "group_name": f"Class {chr(65 + i)}",
            "group_description": "Synthetic generated class",
            "list_vessels": ",\n".join(class_vessels),
            "is_entry": True,
            "is_valid_name": True,
            "is_valid_vessels": True,
        })

    return pd.DataFrame(groups)


def generate_projects(
    num_activities: int,
    locations_df: pd.DataFrame,
    vessel_groups_df: pd.DataFrame,
    vessels_df: pd.DataFrame,
    complexity: str,
    tightness: str,
    overlap_density: float = 0.0,
    multi_vessel_frac: float | None = None,
) -> pd.DataFrame:
    projects = []
    loc_names = locations_df["location_name"].tolist()
    all_vessels = vessels_df["vessel_name"].tolist()
    num_vessels = len(all_vessels)

    group_to_vessels = {}
    for _, row in vessel_groups_df.iterrows():
        g_name = row["group_name"]
        v_list_str = row["list_vessels"]
        if v_list_str == "*":
            group_to_vessels[g_name] = set(all_vessels)
        else:
            # Clean up and split the string "Vessel 1,\nVessel 2"
            v_list = [v.strip() for v in v_list_str.replace('\n', '').split(",") if v.strip()]
            group_to_vessels[g_name] = set(v_list)

    base_date = datetime(2026, 1, 1)

    # --- TIME SLOT ASSIGNMENT ---
    if overlap_density > 0.0:
        n_overlap = round(num_activities * overlap_density)
        n_spaced  = num_activities - n_overlap
        n_groups  = max(1, round(n_overlap / 2.5))  # ~2-3 activities per group

        ov_offsets: list[int] = []
        ov_flags:   list[bool] = []
        for g in range(n_groups):
            g_anchor = random.randint(0, max(0, int(HORIZON_DAYS * 0.60) - DURATION_MAX_DAYS))
            n_in_group = (n_overlap // n_groups) + (1 if g < (n_overlap % n_groups) else 0)
            for _ in range(n_in_group):
                ov_offsets.append(g_anchor + random.randint(0, 20))
                ov_flags.append(True)

        sp_interval = max(1, HORIZON_DAYS // max(1, n_spaced + 1))
        sp_offsets = [
            int((k + 1) * sp_interval + random.uniform(0, sp_interval * 0.4))
            for k in range(n_spaced)
        ]
        sp_flags = [False] * n_spaced

        combined = list(zip(ov_offsets + sp_offsets, ov_flags + sp_flags))
        random.shuffle(combined)
        all_offsets, is_overlap_flags = map(list, zip(*combined))
    else:
        avg_duration = (DURATION_MIN_DAYS + DURATION_MAX_DAYS) / 2
        est_vessel_days_needed = num_activities * avg_duration * 2
        safe_horizon = max(HORIZON_DAYS, int((est_vessel_days_needed / max(1, num_vessels)) * 1.25))
        interval_size = max(1, (safe_horizon - DURATION_MAX_DAYS - WINDOW_LOOSE_MAX_DAYS) // max(1, num_activities))
        all_offsets = [
            int((i * interval_size) + random.uniform(0, interval_size * 0.8))
            for i in range(num_activities)
        ]
        is_overlap_flags = [False] * num_activities

    # --- MULTI-VESSEL FLAGS ---
    if multi_vessel_frac is not None:
        n_multi = round(num_activities * multi_vessel_frac)
        multi_flags: list[bool | None] = [True] * n_multi + [False] * (num_activities - n_multi) # type: ignore
        random.shuffle(multi_flags)
    else:
        multi_flags = [None] * num_activities  # type: ignore[assignment]

    for i, (begin_offset, is_overlap, is_multi) in enumerate(zip(all_offsets, is_overlap_flags, multi_flags)):
        start_loc = random.choice(loc_names)
        remaining = [l for l in loc_names if l != start_loc] or loc_names
        end_loc = random.choice(remaining)

        # Generate dates and duration
        base_start_date = base_date + timedelta(days=int(begin_offset))
        duration = int(random.triangular(DURATION_MIN_DAYS, DURATION_MAX_DAYS, DURATION_MIN_DAYS))

        # Overlap-group activities always use a start window so the scheduling
        # conflict is real and not trivially resolved by a fixed start date.
        if is_overlap:
            has_actual_start = False
        else:
            has_actual_start = random.random() < 0.3

        if has_actual_start:
            actual_start_str = base_start_date.strftime("%d/%m/%Y")
            begin_window_str = ""
            end_window_str = ""
        else:
            actual_start_str = ""
            if tightness == "tight":
                window_width = random.randint(WINDOW_TIGHT_MIN_DAYS, WINDOW_TIGHT_MAX_DAYS)
            else:
                window_width = random.randint(WINDOW_LOOSE_MIN_DAYS, WINDOW_LOOSE_MAX_DAYS)

            e_window = base_start_date + timedelta(days=window_width)
            begin_window_str = base_start_date.strftime("%d/%m/%Y")
            end_window_str = e_window.strftime("%d/%m/%Y")

        # 3. Determine Number of Requirements
        if is_multi is True:
            num_reqs = 2
        elif is_multi is False:
            num_reqs = 1
        else:
            # Backward-compatible: use complexity string
            rand_val = random.random()
            if complexity == "low":
                num_reqs = 1
            elif complexity == "medium":
                num_reqs = 2 if rand_val < 0.30 else 1
            else:
                if rand_val < 0.20:
                    num_reqs = 3
                elif rand_val < 0.50:
                    num_reqs = 2
                else:
                    num_reqs = 1

        # 4. Generate configurations while managing available capacity
        pool = set(all_vessels)
        configs = []
        counts = []
        used_configs = set() # Prevent picking "Class A" twice in the same project

        for _ in range(num_reqs):
            if not pool:
                break # All vessels have been theoretically assigned

            # 30% chance to pick a specific vessel, 70% to pick a group
            pick_specific = random.random() < 0.30

            if pick_specific:
                available_vessels = list(pool)
                if available_vessels:
                    v = random.choice(available_vessels)
                    configs.append(v)
                    counts.append("1")
                    pool.remove(v)
                    used_configs.add(v)
            else:
                # Find groups that still have at least 1 vessel available in the pool
                valid_groups = [g for g, g_v in group_to_vessels.items() if (g_v & pool) and g not in used_configs]
                
                if valid_groups:
                    g = random.choice(valid_groups)
                    avail_in_g = list(group_to_vessels[g] & pool)
                    
                    # Randomly assign between 1 and the maximum available vessels in this group
                    max_k = max(1, int(len(avail_in_g) * 0.5)) 
                    k = random.randint(1, max_k)
                    
                    configs.append(g)
                    counts.append(str(k))
                    used_configs.add(g)
                    
                    # Deduct the 'k' consumed vessels from the project's pool 
                    # so they can't be used by subsequent requirements
                    consumed = random.sample(avail_in_g, k)
                    for cv in consumed:
                        pool.remove(cv)
                else:
                    # Fallback to a specific vessel if groups are exhausted/unavailable
                    available_vessels = list(pool)
                    if available_vessels:
                        v = random.choice(available_vessels)
                        configs.append(v)
                        counts.append("1")
                        pool.remove(v)
                        used_configs.add(v)

        projects.append({
            "row_id": i + 1,
            "project_code": f"PROJ-{str(i+1).zfill(3)}",
            "activity_name": f"Synthetic Operation {i+1}",
            "actual_start": actual_start_str,
            "begin_start_window": begin_window_str,
            "end_start_window": end_window_str,
            "duration_fixed": duration,
            "duration_options": "",
            "duration_opex": "",
            "start_location_name": start_loc,
            "end_location_name": end_loc,
            "route_description": f"{start_loc} - {end_loc}",
            "vessel_configs": ",\n".join(configs),
            "vessel_counts": ",\n".join(counts),
            "is_entry": True,
            "is_valid_name": True,
            "is_valid_dates": True,
            "is_valid_locations": True,
            "is_valid_durations": True,
            "is_valid_configs": True,
            "is_valid_counts": True,
        })

    return pd.DataFrame(projects)


def generate_maintenance(
    num_vessels: int,
    vessels_df: pd.DataFrame,
    locations_df: pd.DataFrame,
) -> pd.DataFrame:
    maint = []
    vessel_names = vessels_df["vessel_name"].tolist()

    num_maint = max(1, round(num_vessels * MAINT_VESSEL_FRACTION))
    selected_vessels = random.sample(vessel_names, k=num_maint)

    base_date = datetime(2026, 3, 1)

    for i, vessel in enumerate(selected_vessels):
        window_offset = (HORIZON_DAYS // num_maint) * i + random.randint(0, 20)
        b_window = base_date + timedelta(days=window_offset)
        e_window = b_window + timedelta(days=45)
        duration = random.randint(MAINT_DURATION_MIN, MAINT_DURATION_MAX)

        maint.append({
            "row_id": i + 1,
            "activity_name": f"Dry Dock {vessel}",
            "actual_start": "",
            "begin_start_window": b_window.strftime("%d/%m/%Y"),
            "end_start_window": e_window.strftime("%d/%m/%Y"),
            "duration": duration,
            "target_vessel": vessel,
            "possible_locations": "All",
            "is_entry": True,
            "is_valid_name": True,
            "is_valid_dates": True,
            "is_valid_durations": True,
            "is_valid_vessel": True,
            "is_valid_locations": True,
        })

    return pd.DataFrame(maint)


def generate_current_locations(
    num_vessels: int,
    cluster: str = DEFAULT_CLUSTER,
) -> list[dict]:
    """
    Generate a current_locations list in the format expected by load_input_json.

    Every entry uses `_vessel_name(i)` — the same function used by
    `generate_vessels()` — so `target_vessel` always matches `vessel_name`
    in the Excel file exactly.

    Each vessel is placed at a random point **within the same geographic
    cluster** as the dataset locations, keeping initial travel times short
    and consistent with the feasibility guard-rails applied elsewhere.

    Parameters
    ----------
    num_vessels : number of vessels (must match the dataset being generated)
    cluster     : geographic cluster — must match the cluster used for
                  generate_locations() in the same dataset

    Returns
    -------
    list of dicts, each with keys:
        is_current_location : bool  (always True)
        target_vessel       : str   (exact vessel name)
        location            : [longitude, latitude]
    """
    lon_range, lat_range, _ = _cluster_coords(cluster)
    entries = []

    for i in range(num_vessels):
        entries.append({
            "is_current_location": True,
            "target_vessel": _vessel_name(i),
            "location": [
                round(random.uniform(*lon_range), 4),
                round(random.uniform(*lat_range), 4),
            ],
        })

    return entries


# =============================================================================
# EXPORTER LOGIC
# =============================================================================

def export_to_excel(filepath: str, dfs: dict) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    for sheet_name, df in dfs.items():
        ws = wb.create_sheet(title=sheet_name)
        meta = METADATA.get(sheet_name)
        if not meta:
            continue

        for r_idx, row_data in enumerate(meta):
            for c_idx, val in enumerate(row_data):
                ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False)):
            for c_idx, value in enumerate(row):
                ws.cell(row=r_idx + 7, column=c_idx + 1, value=value)

    wb.save(filepath)


def export_current_locations(filepath: str, entries: list[dict]) -> None:
    """Write a current_locations list to a JSON file."""
    with open(filepath, "w") as f:
        json.dump(entries, f, indent=4)


# =============================================================================
# MASTER GENERATOR
# =============================================================================

def generate_dataset(
    num_vessels: int,
    num_activities: int,
    num_locations: int,
    complexity: str,
    tightness: str,
    filename: str,
    random_seed: int | None = None,
    cluster: str = DEFAULT_CLUSTER,
    current_locations_path: str | None = None,
    overlap_density: float = 0.0,
    multi_vessel_frac: float | None = None,
) -> None:
    """
    Generate one synthetic HVASP dataset (Excel) and, optionally, its
    companion current_locations JSON file.

    Parameters
    ----------
    num_vessels             : number of vessels
    num_activities          : number of project activities
    num_locations           : number of port/location records (≥ 3 recommended)
    complexity              : "low" | "medium" | "high"
    tightness               : "loose" | "tight"
    filename                : output .xlsx path
    random_seed             : seed for reproducibility; None = fully random
    cluster                 : geographic cluster name (see GEO_CLUSTERS)
    current_locations_path  : if provided, write a companion current_locations
                              JSON file to this path.  The vessel names in the
                              JSON will match the Excel exactly because both
                              generators use _vessel_name().
                              If None, no JSON file is written.
    """
    if random_seed is not None:
        random.seed(random_seed)
        np.random.seed(random_seed)

    df_locations = generate_locations(num_locations, cluster)
    df_vessels = generate_vessels(num_vessels)
    df_vessel_groups = generate_vessel_groups(df_vessels, num_vessels)
    df_projects = generate_projects(
        num_activities, df_locations, df_vessel_groups, df_vessels, complexity, tightness,
        overlap_density=overlap_density, multi_vessel_frac=multi_vessel_frac,
    )
    df_maintenance = generate_maintenance(num_vessels, df_vessels, df_locations)

    dfs = {
        "project_locations": df_locations,
        "maintenance_locations": df_locations.copy(),
        "vessels": df_vessels,
        "vessel_groups": df_vessel_groups,
        "projects": df_projects,
        "maintenance": df_maintenance,
    }

    export_to_excel(filename, dfs)
    print(f"Generated Excel : {filename}  (seed={random_seed}, cluster={cluster})")

    if current_locations_path is not None:
        # Re-seed so current_locations coordinates are deterministic but
        # independent of the main dataset generation order
        if random_seed is not None:
            random.seed(random_seed + 99999)
            np.random.seed(random_seed + 99999)

        entries = generate_current_locations(num_vessels, cluster)
        export_current_locations(current_locations_path, entries)
        print(f"Generated JSON  : {current_locations_path}  ({num_vessels} vessels)")